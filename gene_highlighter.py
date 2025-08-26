#!/usr/bin/env python3
"""
遺伝子・タンパク質名認識とExcelハイライトツール
scispaCyを使用して生物医学文書から遺伝子・タンパク質名を自動認識し、
Excelファイルでハイライト表示する
"""

import pandas as pd
import spacy
import scispacy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from typing import List, Tuple, Dict
import argparse

class GeneProteinHighlighter:
    def __init__(self, model_name: str = "en_ner_bionlp13cg_md"):
        """
        遺伝子・タンパク質名ハイライトツールの初期化
        
        Args:
            model_name: 使用するscispaCyモデル名
                      - en_ner_bionlp13cg_md: 遺伝子認識用
                      - en_ner_jnlpba_md: タンパク質認識用
                      - en_ner_bc5cdr_md: 化学物質・疾患認識用
        """
        try:
            self.nlp = spacy.load(model_name)
            self.model_name = model_name
            print(f"Successfully loaded {model_name} model")
        except IOError:
            print(f"Error: {model_name} model not found")
            print("Installation:")
            print(f"pip install https://s3-us-west-2.amazonaws.com/ai2-s2-scispacy/releases/v0.5.4/{model_name}-0.5.4.tar.gz")
            raise
        
        # ハイライト色の設定
        self.highlight_colors = {
            'GENE_OR_GENE_PRODUCT': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # 黄色
            'PROTEIN': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # ライトグリーン
            'CHEMICAL': PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),  # ライトサーモン
            'DISEASE': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # ライトピンク
        }
        
    def extract_entities(self, text: str) -> List[Tuple[str, str, int, int]]:
        """
        テキストから生物医学エンティティを抽出
        
        Args:
            text: 解析するテキスト
            
        Returns:
            List of (entity_text, label, start_pos, end_pos)
        """
        if pd.isna(text) or not isinstance(text, str):
            return []
        
        doc = self.nlp(text)
        entities = []
        
        for ent in doc.ents:
            # 関心のあるエンティティタイプのみフィルタ
            if ent.label_ in ['GENE_OR_GENE_PRODUCT', 'PROTEIN', 'CHEMICAL', 'DISEASE']:
                entities.append((ent.text, ent.label_, ent.start_char, ent.end_char))
        
        return entities
    
    def process_excel_file(self, input_file: str, output_file: str, 
                          columns: List[str] = None, sheet_name: str = None):
        """
        Excelファイルを処理して遺伝子・タンパク質名をハイライト
        
        Args:
            input_file: 入力Excelファイルパス
            output_file: 出力Excelファイルパス  
            columns: 処理対象の列名リスト（Noneの場合は全列）
            sheet_name: 処理するシート名（Noneの場合は最初のシート）
        """
        try:
            # Excelファイルを読み込み
            if sheet_name:
                df = pd.read_excel(input_file, sheet_name=sheet_name)
            else:
                df = pd.read_excel(input_file)
            print(f"Loaded {input_file}")
            print(f"  - Rows: {len(df)}")
            print(f"  - Columns: {len(df.columns)}")
            
            # 処理対象列を決定
            if columns is None:
                target_columns = df.columns.tolist()
            else:
                target_columns = [col for col in columns if col in df.columns]
            
            print(f"  - Target columns: {target_columns}")
            
            # エンティティ抽出とレポート生成
            entity_report = {}
            total_entities = 0
            
            for col in target_columns:
                col_entities = []
                for idx, cell_value in df[col].items():
                    if pd.notna(cell_value):
                        entities = self.extract_entities(str(cell_value))
                        for entity_text, entity_label, start_pos, end_pos in entities:
                            col_entities.append({
                                'row': idx + 2,  # Excelの行番号（1-indexed + header）
                                'text': entity_text,
                                'label': entity_label,
                                'full_text': str(cell_value)
                            })
                            total_entities += 1
                
                entity_report[col] = col_entities
            
            print(f"\nFound entities: {total_entities}")
            
            # レポートの表示
            for col, entities in entity_report.items():
                if entities:
                    print(f"\nColumn '{col}' results:")
                    entity_counts = {}
                    for ent in entities:
                        label = ent['label']
                        entity_counts[label] = entity_counts.get(label, 0) + 1
                    
                    for label, count in entity_counts.items():
                        print(f"   {label}: {count}")
            
            # Excelファイルにハイライト適用
            self._apply_highlighting(input_file, output_file, entity_report, 
                                   target_columns, sheet_name)
            
        except Exception as e:
            print(f"Error: {e}")
            raise
    
    def _apply_highlighting(self, input_file: str, output_file: str, 
                           entity_report: Dict, target_columns: List[str], 
                           sheet_name: str = None):
        """
        Excelファイルにハイライトを適用
        """
        # openpyxlでワークブックを開く
        wb = load_workbook(input_file)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # 列のインデックスマップを作成
        header_row = 1
        col_index_map = {}
        for col_num, cell in enumerate(ws[header_row], 1):
            if cell.value in target_columns:
                col_index_map[cell.value] = col_num
        
        # ハイライトを適用
        highlighted_cells = 0
        for col_name, entities in entity_report.items():
            if col_name in col_index_map:
                col_index = col_index_map[col_name]
                for entity in entities:
                    row_num = entity['row']
                    entity_label = entity['label']
                    
                    if entity_label in self.highlight_colors:
                        cell = ws.cell(row=row_num, column=col_index)
                        cell.fill = self.highlight_colors[entity_label]
                        highlighted_cells += 1
        
        # ファイル保存
        wb.save(output_file)
        print(f"Highlighted {highlighted_cells} cells")
        print(f"Saved results to {output_file}")
        
        # 凡例を追加
        self._add_legend(wb, ws)
        wb.save(output_file)
    
    def _add_legend(self, wb, ws):
        """ハイライト凡例をワークシートに追加"""
        # 凡例の開始位置を決定
        last_row = ws.max_row
        legend_start_row = last_row + 3
        
        ws.cell(row=legend_start_row, column=1, value="Highlight Legend:")
        ws.cell(row=legend_start_row, column=1).font = ws.cell(row=legend_start_row, column=1).font.copy(bold=True)
        
        legend_items = {
            'GENE_OR_GENE_PRODUCT': ('Gene/Gene Product', self.highlight_colors['GENE_OR_GENE_PRODUCT']),
            'PROTEIN': ('Protein', self.highlight_colors['PROTEIN']),
            'CHEMICAL': ('Chemical', self.highlight_colors['CHEMICAL']),
            'DISEASE': ('Disease', self.highlight_colors['DISEASE'])
        }
        
        for i, (label, (description, fill)) in enumerate(legend_items.items()):
            row = legend_start_row + 1 + i
            ws.cell(row=row, column=1, value=description)
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=2, value=f"({label})")

def main():
    """メイン関数 - コマンドライン実行用"""
    parser = argparse.ArgumentParser(description='Gene/Protein Highlighter Tool')
    parser.add_argument('input_file', help='Input Excel file path')
    parser.add_argument('-o', '--output', help='Output Excel file path (auto-generated if omitted)')
    parser.add_argument('-c', '--columns', nargs='+', help='Target column names')
    parser.add_argument('-s', '--sheet', help='Sheet name to process')
    parser.add_argument('-m', '--model', default='en_ner_bionlp13cg_md', 
                       help='scispaCy model name to use')
    
    args = parser.parse_args()
    
    # 出力ファイル名の自動生成
    if not args.output:
        base_name = args.input_file.rsplit('.', 1)[0]
        args.output = f"{base_name}_highlighted.xlsx"
    
    # ツールの実行
    highlighter = GeneProteinHighlighter(args.model)
    highlighter.process_excel_file(
        input_file=args.input_file,
        output_file=args.output,
        columns=args.columns,
        sheet_name=args.sheet
    )

if __name__ == "__main__":
    # 使用例
    print("Gene/Protein Highlighter Tool")
    print("\nRequired dependencies:")
    print("pip install pandas openpyxl spacy scispacy")
    print("pip install https://s3-us-west-2.amazonaws.com/ai2-s2-scispacy/releases/v0.5.4/en_ner_bionlp13cg_md-0.5.4.tar.gz")
    
    print("\nUsage:")
    print("python gene_highlighter.py input.xlsx -c 'Abstract' 'Title' -o output.xlsx")
    
    # サンプルテスト（コマンドライン引数がない場合）
    import sys
    import os
    if len(sys.argv) == 1:
        print("\n" + "="*60)
        print("DEMO: Gene/Protein Recognition Test")
        print("="*60)
        
        try:
            highlighter = GeneProteinHighlighter()
            
            # サンプルテキストでのテスト
            sample_texts = [
                "BRCA1 mutations are associated with p53 pathway disruption in breast cancer.",
                "The EGFR protein shows overexpression in lung cancer patients treated with erlotinib.",
                "TP53 mutations lead to loss of DNA damage response and increased cancer risk.",
                "MYC oncogene amplification drives tumor cell proliferation in various cancers.",
                "Alzheimer's disease is characterized by amyloid beta plaques and tau protein aggregation."
            ]
            
            print("\n1. Text Analysis Results:")
            print("-" * 40)
            
            total_entities = 0
            all_entities = {}
            
            for i, text in enumerate(sample_texts, 1):
                entities = highlighter.extract_entities(text)
                total_entities += len(entities)
                
                print(f"\nSample {i}:")
                print(f"Text: {text}")
                print(f"Found: {len(entities)} entities")
                
                for entity_text, label, start, end in entities:
                    print(f"  -> '{entity_text}' [{label}] (pos: {start}-{end})")
                    if label not in all_entities:
                        all_entities[label] = []
                    all_entities[label].append(entity_text)
            
            print(f"\n2. Summary Statistics:")
            print("-" * 40)
            print(f"Total entities found: {total_entities}")
            for label, entities_list in all_entities.items():
                unique_entities = list(set(entities_list))
                print(f"{label}: {len(entities_list)} total, {len(unique_entities)} unique")
                print(f"  Examples: {', '.join(unique_entities[:3])}")
            
            # Excelファイルが存在する場合のデモ
            if os.path.exists('sample_data.xlsx'):
                print(f"\n3. Excel File Demo:")
                print("-" * 40)
                print("Found sample_data.xlsx - Analyzing content...")
                
                try:
                    df = pd.read_excel('sample_data.xlsx')
                    print(f"[OK] File contains {len(df)} rows and {len(df.columns)} columns")
                    
                    # 各列のエンティティを分析
                    excel_total_entities = 0
                    for col in df.columns:
                        col_entities = 0
                        for cell_value in df[col]:
                            if pd.notna(cell_value):
                                entities = highlighter.extract_entities(str(cell_value))
                                col_entities += len(entities)
                        excel_total_entities += col_entities
                        print(f"  Column '{col}': {col_entities} entities found")
                    
                    print(f"[OK] Total entities in Excel: {excel_total_entities}")
                    print("[INFO] Run 'python gene_highlighter.py sample_data.xlsx' for full processing")
                    
                except Exception as e:
                    print(f"Could not analyze Excel file: {e}")
                    
            else:
                print(f"\n3. Excel Demo:")
                print("-" * 40)
                print("To test with Excel files:")
                print("1. Create an Excel file with biomedical text")
                print("2. Run: python gene_highlighter.py your_file.xlsx")
                print("3. Check the highlighted output file")
            
            print(f"\n4. Available Entity Types:")
            print("-" * 40)
            color_info = {
                'GENE_OR_GENE_PRODUCT': 'Yellow - Genes and gene products',
                'PROTEIN': 'Green - Proteins',
                'CHEMICAL': 'Orange - Chemical compounds',
                'DISEASE': 'Pink - Diseases and conditions'
            }
            
            for entity_type, description in color_info.items():
                print(f"  {description}")
            
            print(f"\n" + "="*60)
            print("Demo completed! Try with your own Excel files:")
            print("python gene_highlighter.py your_file.xlsx")
            print("="*60)
            
        except Exception as e:
            print(f"Error during demo: {e}")
            print("Make sure all dependencies are installed:")
            print("pip install -r requirements.txt")
    else:
        main()